"""
Bulk Excel Part Number Search Parser
Handles user-uploaded Excel files with standardized part number search format
"""

from __future__ import annotations

import io
import pandas as pd
import re
from typing import Union
try:
    import numpy as np
except Exception:  # pragma: no cover
    np = None  # type: ignore
from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass
from enum import Enum

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class MatchStatus(Enum):
    FOUND = "found"
    PARTIAL = "partial"
    NOT_FOUND = "not_found"


@dataclass
class UserPartData:
    """Standardized user part data from Excel upload"""
    part_number: str
    part_name: str = ""
    quantity: int = 0
    manufacturer_name: str = ""
    row_index: int = 0  # Original row number for error reporting


@dataclass
class BulkSearchConfig:
    """Configuration for bulk search processing"""
    max_file_size_mb: int = 50
    batch_size: int = 500
    max_results_per_part: int = 3
    required_headers: List[str] = None
    processing_timeout_seconds: int = 30
    enable_manufacturer_cross_check: bool = True
    confidence_threshold: float = 0.3
    
    def __post_init__(self):
        if self.required_headers is None:
            self.required_headers = ["Part Number", "Part name", "Quantity", "Manufacturer name"]


class BulkExcelParser:
    """Parser for bulk Excel part number search files"""
    
    def __init__(self, config: BulkSearchConfig = None):
        self.config = config or BulkSearchConfig()
        
    def validate_headers(self, headers: List[str]) -> Tuple[bool, str, Dict[str, str]]:
        """
        Validate and map Excel headers to standard format.
        Returns: (is_valid, error_message, column_mapping)
        """
        if not headers:
            return False, "No headers found in file", {}
            
        # Normalize headers (trim, case-insensitive)
        normalized_headers = [h.strip().lower() for h in headers]
        required_lower = [h.lower() for h in self.config.required_headers]
        
        # Create mapping from normalized to original headers
        header_mapping = {}
        for orig, norm in zip(headers, normalized_headers):
            header_mapping[norm] = orig
            
        # Check for required headers with flexible matching
        missing_headers = []
        column_mapping = {}
        
        # Flexible header matching - only check for required headers from config
        header_variations = {
            "part number": ["part number", "part_number", "partnumber", "part no", "partno", "pn"],
            "part name": ["part name", "part_name", "partname", "description", "desc", "item name"],
            "quantity": ["quantity", "qty", "amount", "count", "units"],
            "manufacturer name": ["manufacturer name", "manufacturer_name", "manufacturer", "mfg", "brand", "supplier"]
        }
        
        # Only check for headers that are required in the config
        for required_field in self.config.required_headers:
            required_field_lower = required_field.lower()
            if required_field_lower in header_variations:
                variations = header_variations[required_field_lower]
                found = False
                for variation in variations:
                    if variation in normalized_headers:
                        column_mapping[required_field_lower] = header_mapping[variation]
                        found = True
                        break
                if not found:
                    missing_headers.append(required_field_lower)
        
        if missing_headers:
            return False, f"Missing required headers: {missing_headers}. Found: {list(header_mapping.keys())}", {}
            
        return True, "", column_mapping
    
    def parse_excel_file(self, file_bytes: bytes, filename: str) -> Tuple[List[UserPartData], List[str]]:
        """
        Parse Excel file and extract user part data.
        Returns: (user_parts, error_messages)
        """
        errors = []
        user_parts = []
        
        try:
            # Detect file format
            if filename.lower().endswith('.csv'):
                df = pd.read_csv(io.BytesIO(file_bytes))
            else:
                # Excel file
                if load_workbook is None:
                    df = pd.read_excel(io.BytesIO(file_bytes), engine='openpyxl')
                else:
                    # Use openpyxl for better performance on large files and robust header detection
                    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
                    ws = wb.worksheets[0]

                    # Extract up to first 2000 rows defensively
                    raw_rows = []
                    max_take = 2000
                    count = 0
                    for row in ws.iter_rows(values_only=True):
                        raw_rows.append(list(row))
                        count += 1
                        if count >= max_take:
                            break
                    wb.close()

                    if not raw_rows:
                        return [], ["File is empty"]

                    # Find header row by matching required header variations
                    header_variations = {
                        "part number": ["part number", "part_number", "partnumber", "part no", "partno", "pn"],
                        "part name": ["part name", "part_name", "partname", "description", "desc", "item name"],
                        "quantity": ["quantity", "qty", "amount", "count", "units"],
                        "manufacturer name": ["manufacturer name", "manufacturer_name", "manufacturer", "mfg", "brand", "supplier"],
                    }

                    def normalize_cell(v):
                        if v is None:
                            return ""
                        return str(v).strip().lower()

                    header_row_index = 0
                    best_score = -1
                    # Scan first 20 rows for a plausible header row
                    for idx, row in enumerate(raw_rows[:20]):
                        normalized = [normalize_cell(c) for c in row]
                        score = 0
                        for required, variants in header_variations.items():
                            if any(v in normalized for v in variants):
                                score += 1
                        if score > best_score:
                            best_score = score
                            header_row_index = idx

                    headers_raw = raw_rows[header_row_index]
                    # Determine width using the widest row among header+next 50 rows
                    width = max(len(headers_raw), max((len(r) for r in raw_rows[header_row_index:header_row_index+50]), default=len(headers_raw)))

                    # Build headers with fallbacks "Column_i" for empty cells
                    headers = []
                    for i in range(width):
                        val = headers_raw[i] if i < len(headers_raw) else None
                        headers.append(str(val) if val is not None and str(val).strip() else f"Column_{i}")

                    # Collect data rows after the header row, pad/truncate to width
                    data_rows = []
                    for r in raw_rows[header_row_index+1:]:
                        row_vals = list(r)
                        if len(row_vals) < width:
                            row_vals = row_vals + [None] * (width - len(row_vals))
                        elif len(row_vals) > width:
                            row_vals = row_vals[:width]
                        data_rows.append(row_vals)

                    # Drop leading completely empty rows
                    def is_all_empty(row):
                        return all((c is None or str(c).strip() == "") for c in row)
                    while data_rows and is_all_empty(data_rows[0]):
                        data_rows.pop(0)

                    if not data_rows:
                        return [], ["No data rows found after header"]

                    df = pd.DataFrame(data_rows, columns=headers)
            
            if df.empty:
                return [], ["File contains no data"]
            
            # Validate headers
            headers = list(df.columns)
            is_valid, error_msg, column_mapping = self.validate_headers(headers)
            
            if not is_valid:
                return [], [error_msg]
            
            # Helper to normalize part number values (e.g., 3585720.0 -> "3585720")
            def normalize_part_number_value(value: Any) -> str:
                # Direct string handling first
                if isinstance(value, str):
                    s = value.replace('\u00A0', ' ').replace(',', '').strip()
                    # drop trailing .0 or .00... if the rest are digits
                    if re.fullmatch(r"\d+\.0+", s):
                        return s.split(".")[0]
                    if re.fullmatch(r"\d+", s):
                        return s
                    # Also handle strings like '3585720.00 '
                    m = re.fullmatch(r"(\d+)\.(0+)", s)
                    if m:
                        return m.group(1)
                    return s
                # Numeric types: numpy or python
                try:
                    # numpy numeric types
                    if np is not None and isinstance(value, (np.integer, np.floating)):
                        # If float but integral, cast to int then str
                        try:
                            f = float(value)
                            if float(f).is_integer():
                                return str(int(f))
                            return str(value)
                        except Exception:
                            return str(value)
                except Exception:
                    pass
                if isinstance(value, (int,)):
                    return str(int(value))
                if isinstance(value, float):
                    if float(value).is_integer():
                        return str(int(value))
                    return str(value)
                # Fallback
                return str(value).replace('\u00A0', ' ').replace(',', '').strip()

            # Process each row
            for idx, row in df.iterrows():
                try:
                    # Extract data using column mapping - only access columns that exist
                    raw_pn = row[column_mapping["part number"]] if pd.notna(row[column_mapping["part number"]]) else ""
                    part_number = normalize_part_number_value(raw_pn)
                    
                    # Extract optional fields only if they exist in column mapping
                    part_name = ""
                    if "part name" in column_mapping:
                        part_name = str(row[column_mapping["part name"]]).strip() if pd.notna(row[column_mapping["part name"]]) else ""
                    
                    manufacturer_name = ""
                    if "manufacturer name" in column_mapping:
                        manufacturer_name = str(row[column_mapping["manufacturer name"]]).strip() if pd.notna(row[column_mapping["manufacturer name"]]) else ""
                    
                    # Parse quantity only if it exists
                    quantity = 0
                    if "quantity" in column_mapping:
                        quantity_raw = row[column_mapping["quantity"]]
                        if pd.isna(quantity_raw):
                            quantity = 0
                        else:
                            try:
                                # Handle various quantity formats
                                if isinstance(quantity_raw, str):
                                    quantity_raw = quantity_raw.replace(',', '').strip()
                                quantity = int(float(quantity_raw))
                            except (ValueError, TypeError):
                                quantity = 0
                                errors.append(f"Row {idx + 2}: Invalid quantity '{quantity_raw}', using 0")
                    
                    # Skip rows with empty part number
                    if not part_number or part_number.lower() in ['nan', 'none', '']:
                        errors.append(f"Row {idx + 2}: Empty part number, skipping")
                        continue
                    
                    user_part = UserPartData(
                        part_number=part_number,
                        part_name=part_name,
                        quantity=quantity,
                        manufacturer_name=manufacturer_name,
                        row_index=idx + 2  # +2 for 1-based indexing and header row
                    )
                    user_parts.append(user_part)
                    
                except Exception as e:
                    errors.append(f"Row {idx + 2}: Error processing row - {str(e)}")
                    continue
            
            # Limit to reasonable size
            if len(user_parts) > 50000:  # 50K parts max
                user_parts = user_parts[:50000]
                errors.append(f"File contains {len(user_parts)} parts, limited to 50,000 for performance")
            
            return user_parts, errors
            
        except Exception as e:
            return [], [f"Failed to parse file: {str(e)}"]
    
    def validate_file_size(self, file_bytes: bytes) -> Tuple[bool, str]:
        """Validate file size is within limits"""
        size_mb = len(file_bytes) / (1024 * 1024)
        if size_mb > self.config.max_file_size_mb:
            return False, f"File size {size_mb:.1f}MB exceeds limit of {self.config.max_file_size_mb}MB"
        return True, ""
