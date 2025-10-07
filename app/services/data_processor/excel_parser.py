from typing import Iterable, Iterator, List
from io import BytesIO
import csv
import pandas as pd
from app.services.data_processor.schema_def import validate_headers, expected_headers
from app.core.config import settings

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None  # Fallback handled at runtime


def detect_format(filename: str) -> str:
    name = filename.lower()
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return "xlsx"
    return "csv"


def iter_rows(file_bytes: bytes, filename: str, chunk_size: int = 10000, skip_rows: int = 0) -> Iterator[List[dict]]:
    fmt = detect_format(filename)
    if fmt == "csv":
        text_stream = BytesIO(file_bytes)
        # First, read header to validate
        text_stream.seek(0)
        sample = pd.read_csv(text_stream, nrows=0)
        ok, msg = validate_headers([str(c).strip() for c in list(sample.columns)])
        if not ok:
            raise ValueError(msg)
        text_stream.seek(0)
        skipped = 0
        for chunk in pd.read_csv(text_stream, chunksize=chunk_size):
            records = chunk.where(pd.notnull(chunk), None).to_dict(orient="records")
            if skip_rows and skipped < skip_rows:
                if skipped + len(records) <= skip_rows:
                    skipped += len(records)
                    continue
                # drop first part
                drop = skip_rows - skipped
                records = records[drop:]
                skipped = skip_rows
            if records:
                yield records
    else:
        # Stream XLSX using openpyxl to avoid loading entire sheet
        if load_workbook is None:
            # Fallback to pandas full read if openpyxl is unavailable
            bio = BytesIO(file_bytes)
            try:
                # Use ExcelFile to iterate all sheets
                xls = pd.ExcelFile(bio, engine="openpyxl")
                total_skipped = 0
                for sheet_name in xls.sheet_names:
                    df = xls.parse(sheet_name)
                    if df is None or df.empty:
                        continue
                    ok, msg = validate_headers([str(c).strip() for c in list(df.columns)])
                    if not ok:
                        # Skip sheets that don't match expected schema
                        continue
                    cols = [c for c in df.columns if c in expected_headers()]
                    df = df[cols]
                    df = df.where(pd.notnull(df), None)
                    records = df.to_dict(orient="records")
                    # Apply cross-sheet skipping
                    if skip_rows and total_skipped < skip_rows:
                        take_from_skip = min(len(records), skip_rows - total_skipped)
                        records = records[take_from_skip:]
                        total_skipped += take_from_skip
                    for start in range(0, len(records), chunk_size):
                        chunk = records[start:start + chunk_size]
                        if chunk:
                            yield chunk
                return
            except Exception:
                # Fall back to single-sheet read
                bio.seek(0)
                sample = pd.read_excel(bio, engine="openpyxl", nrows=0)
                ok, msg = validate_headers([str(c).strip() for c in list(sample.columns)])
                if not ok:
                    raise ValueError(msg)
                bio.seek(0)
                df = pd.read_excel(bio, engine="openpyxl")
                cols = [c for c in df.columns if c in expected_headers()]
                df = df[cols]
                df = df.where(pd.notnull(df), None)
                records = df.to_dict(orient="records")
                for start in range(0, len(records), chunk_size):
                    yield records[start:start + chunk_size]
                return

        bio = BytesIO(file_bytes)
        wb = load_workbook(filename=bio, read_only=True, data_only=True)
        total_skipped = 0
        batch: List[dict] = []

        try:
            for ws in wb.worksheets:
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header = next(rows_iter)
                except StopIteration:
                    continue
                header_list = [str(h).strip() if h is not None else "" for h in header]
                ok, msg = validate_headers(header_list)
                if not ok:
                    # Skip sheets that don't match expected schema
                    continue
                # Maintain only expected columns
                header_to_keep = expected_headers()
                keep_indices = [i for i, h in enumerate(header_list) if h in header_to_keep]

                # Skip already processed data rows (after header) across sheets
                if skip_rows and total_skipped < skip_rows:
                    to_skip_here = skip_rows - total_skipped
                    skipped_here = 0
                    while skipped_here < to_skip_here:
                        try:
                            next(rows_iter)
                        except StopIteration:
                            break
                        skipped_here += 1
                    total_skipped += skipped_here

                for row in rows_iter:
                    if row is None:
                        continue
                    record = {}
                    for idx in keep_indices:
                        col_name = header_list[idx]
                        value = row[idx] if idx < len(row) else None
                        record[col_name] = value
                    batch.append(record)
                    if len(batch) >= chunk_size:
                        yield batch
                        batch = []
            if batch:
                yield batch
        finally:
            wb.close()


